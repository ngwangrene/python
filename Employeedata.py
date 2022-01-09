import openpyxl as xl
import csv
wb =xl.load_workbook('Employeedata.xlsx')
sheet=wb['Sheet1']
oldemail = 'helpinghands.cm'
newemail = 'handsinhands.org'
for i in range (2, sheet.max_row+1):
    cell=sheet.cell(i,2)
    if oldemail in cell.value:
        updated_Email=(cell.value).replace(oldemail,newemail)
        sheet.cell(i,2).value=updated_Email
wb.save('Employeedataupdated.xlsx')
with open('Employeedata.csv', newline='') as myFile:
    reader = csv.reader(myFile)
    myFile = ''.join([i for i in myFile])
    myFile = myFile.replace('@helpinghands.cm', '@handsinhands.org')

    output = open('Employeedataupdated.csv', 'w')

    output.writelines(myFile)
    output.close